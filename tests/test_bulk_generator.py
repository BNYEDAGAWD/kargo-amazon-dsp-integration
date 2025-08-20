"""Tests for bulk sheet generator service."""
import os
import tempfile
import pytest
from datetime import datetime, timedelta
from unittest.mock import AsyncMock, MagicMock

from app.services.bulk_generator import (
    BulkSheetGenerator,
    BulkSheetRequest,
    BulkSheetResponse
)
from app.models.database import CampaignDB, ProcessedCreativeDB, CampaignCreativeAssociationDB


@pytest.fixture
def mock_db_session():
    """Mock database session."""
    session = AsyncMock()
    session.commit = AsyncMock()
    session.execute = AsyncMock()
    return session


@pytest.fixture
def temp_output_dir():
    """Temporary output directory for bulk sheets."""
    with tempfile.TemporaryDirectory() as temp_dir:
        yield temp_dir


@pytest.fixture
def bulk_generator(mock_db_session, temp_output_dir):
    """Bulk sheet generator instance."""
    return BulkSheetGenerator(mock_db_session, temp_output_dir)


@pytest.fixture
def sample_campaign():
    """Sample campaign for bulk sheet generation."""
    return CampaignDB(
        campaign_id="camp_123",
        name="Test Campaign",
        advertiser_id="adv_123",
        status="active",
        phase="phase_1",
        config={
            "targeting": {
                "audiences": ["audience_1", "audience_2"],
                "keywords": ["keyword_1", "keyword_2"]
            }
        },
        viewability_config={"phase": "phase_1", "vendors": ["double_verify"]},
        total_budget=50000.0,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=30),
        creative_count=3,
        processed_creatives_count=3,
        order_id="order_123"
    )


@pytest.fixture
def sample_creatives():
    """Sample processed creatives."""
    return [
        ProcessedCreativeDB(
            creative_id="creative_1",
            name="Display Creative 1",
            format="runway_display",
            original_snippet_url="https://example.com/snippet1",
            processed_code="<div>Display HTML</div>",
            amazon_dsp_ready=True,
            creative_type="display",
            viewability_config={"phase": "phase_1", "vendors": ["double_verify"]},
            processing_metadata={"width": 300, "height": 250, "dimensions": "300x250"},
            original_config={},
            status="processed",
            amazon_creative_id="amz_creative_1"
        ),
        ProcessedCreativeDB(
            creative_id="creative_2",
            name="Video Creative 1",
            format="enhanced_pre_roll_video",
            original_snippet_url="https://example.com/snippet2",
            processed_code="<VAST>Video XML</VAST>",
            amazon_dsp_ready=True,
            creative_type="video",
            viewability_config={"phase": "phase_1", "vendors": ["double_verify"]},
            processing_metadata={"width": 640, "height": 480, "dimensions": "640x480"},
            original_config={},
            status="processed",
            amazon_creative_id="amz_creative_2"
        ),
        ProcessedCreativeDB(
            creative_id="creative_3",
            name="Display Creative 2",
            format="runway_display",
            original_snippet_url="https://example.com/snippet3",
            processed_code="<div>Another Display HTML</div>",
            amazon_dsp_ready=True,
            creative_type="display",
            viewability_config={"phase": "phase_1", "vendors": ["double_verify"]},
            processing_metadata={"width": 728, "height": 90, "dimensions": "728x90"},
            original_config={},
            status="processed"
        )
    ]


@pytest.fixture
def sample_associations(sample_campaign, sample_creatives):
    """Sample campaign-creative associations."""
    return [
        CampaignCreativeAssociationDB(
            campaign_id=sample_campaign.campaign_id,
            creative_id=creative.creative_id,
            line_item_name=f"{sample_campaign.name}_line_item_{creative.format}",
            line_item_type="display" if "display" in creative.format else "video",
            bid=2.50,
            budget=sample_campaign.total_budget / 3,
            status="active"
        )
        for creative in sample_creatives
    ]


class TestBulkSheetGenerator:
    """Test bulk sheet generator functionality."""
    
    @pytest.mark.asyncio
    async def test_generate_excel_bulk_sheet(
        self,
        bulk_generator,
        sample_campaign,
        sample_creatives,
        sample_associations,
        mock_db_session
    ):
        """Test Excel bulk sheet generation."""
        # Mock database responses
        mock_db_session.execute.side_effect = [
            # Campaign query
            MagicMock(scalar_one_or_none=MagicMock(return_value=sample_campaign)),
            # Associations query
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=sample_associations)))),
            # Creatives query
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=sample_creatives))))
        ]
        
        # Generate bulk sheet request
        request = BulkSheetRequest(
            campaign_id="camp_123",
            include_creatives=True,
            include_line_items=True,
            include_targeting=True,
            format="xlsx"
        )
        
        # Generate bulk sheet
        response = await bulk_generator.generate_bulk_sheet(request)
        
        # Assertions
        assert isinstance(response, BulkSheetResponse)
        assert response.file_name.endswith(".xlsx")
        assert "Campaign_Info" in response.sheets
        assert "Creatives" in response.sheets
        assert "Line_Items" in response.sheets
        assert "Targeting" in response.sheets
        assert response.total_rows > 0
        
        # Verify file was created
        assert os.path.exists(response.file_path)
        
        # Verify campaign was updated with bulk sheet path
        assert sample_campaign.bulk_sheet_path == response.file_path
        mock_db_session.commit.assert_called()
    
    @pytest.mark.asyncio
    async def test_generate_csv_bulk_sheet(
        self,
        bulk_generator,
        sample_campaign,
        sample_creatives,
        sample_associations,
        mock_db_session
    ):
        """Test CSV bulk sheet generation."""
        # Mock database responses
        mock_db_session.execute.side_effect = [
            # Campaign query
            MagicMock(scalar_one_or_none=MagicMock(return_value=sample_campaign)),
            # Associations query
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=sample_associations)))),
            # Creatives query
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=sample_creatives))))
        ]
        
        # Generate bulk sheet request
        request = BulkSheetRequest(
            campaign_id="camp_123",
            include_creatives=True,
            include_line_items=True,
            include_targeting=False,
            format="csv"
        )
        
        # Generate bulk sheet
        response = await bulk_generator.generate_bulk_sheet(request)
        
        # Assertions
        assert isinstance(response, BulkSheetResponse)
        assert not response.file_name.endswith(".csv")  # Base name without extension
        assert "campaign_info" in response.sheets
        assert "creatives" in response.sheets
        assert "line_items" in response.sheets
        assert "targeting" not in response.sheets  # Not included
        assert response.total_rows > 0
        
        # Verify CSV files were created
        base_path = response.file_path
        assert os.path.exists(f"{base_path}_campaign_info.csv")
        assert os.path.exists(f"{base_path}_creatives.csv")
        assert os.path.exists(f"{base_path}_line_items.csv")
    
    @pytest.mark.asyncio
    async def test_generate_bulk_sheet_campaign_not_found(
        self,
        bulk_generator,
        mock_db_session
    ):
        """Test bulk sheet generation with missing campaign."""
        # Mock empty campaign result
        mock_db_session.execute.return_value.scalar_one_or_none.return_value = None
        
        # Generate bulk sheet request
        request = BulkSheetRequest(campaign_id="nonexistent")
        
        # Should raise ValueError
        with pytest.raises(ValueError, match="Campaign not found"):
            await bulk_generator.generate_bulk_sheet(request)
    
    @pytest.mark.asyncio
    async def test_minimal_bulk_sheet_generation(
        self,
        bulk_generator,
        sample_campaign,
        mock_db_session
    ):
        """Test minimal bulk sheet with only campaign info."""
        # Mock database responses with empty associations/creatives
        mock_db_session.execute.side_effect = [
            # Campaign query
            MagicMock(scalar_one_or_none=MagicMock(return_value=sample_campaign)),
            # Associations query (empty)
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=[])))),
            # Creatives query (empty)
            MagicMock(scalars=MagicMock(return_value=MagicMock(all=MagicMock(return_value=[]))))
        ]
        
        # Generate minimal bulk sheet
        request = BulkSheetRequest(
            campaign_id="camp_123",
            include_creatives=False,
            include_line_items=False,
            include_targeting=False,
            format="xlsx"
        )
        
        # Generate bulk sheet
        response = await bulk_generator.generate_bulk_sheet(request)
        
        # Assertions
        assert isinstance(response, BulkSheetResponse)
        assert response.sheets == ["Campaign_Info"]
        assert response.total_rows > 0  # At least campaign info rows
        assert os.path.exists(response.file_path)
    
    def test_create_campaign_info_sheet(
        self,
        bulk_generator,
        sample_campaign
    ):
        """Test campaign info sheet creation."""
        # Import openpyxl for testing
        from openpyxl import Workbook
        
        wb = Workbook()
        sheet = wb.active
        
        # Create campaign info sheet
        row_count = bulk_generator._create_campaign_info_sheet(sheet, sample_campaign)
        
        # Assertions
        assert row_count > 1  # Header + data rows
        assert sheet.cell(row=1, column=1).value == "Field"
        assert sheet.cell(row=1, column=2).value == "Value"
        assert sheet.cell(row=2, column=1).value == "Campaign ID"
        assert sheet.cell(row=2, column=2).value == "camp_123"
    
    def test_create_creatives_sheet(
        self,
        bulk_generator,
        sample_creatives
    ):
        """Test creatives sheet creation."""
        from openpyxl import Workbook
        
        wb = Workbook()
        sheet = wb.active
        
        # Create creatives sheet
        row_count = bulk_generator._create_creatives_sheet(
            sheet,
            sample_creatives,
            "phase_1"
        )
        
        # Assertions
        assert row_count == len(sample_creatives) + 1  # Header + creative rows
        assert sheet.cell(row=1, column=1).value == "Creative ID"
        assert sheet.cell(row=2, column=1).value == "creative_1"
        assert sheet.cell(row=2, column=2).value == "Display Creative 1"
        assert sheet.cell(row=2, column=6).value == "phase_1"
    
    def test_create_line_items_sheet(
        self,
        bulk_generator,
        sample_associations,
        sample_campaign
    ):
        """Test line items sheet creation."""
        from openpyxl import Workbook
        
        wb = Workbook()
        sheet = wb.active
        
        # Create line items sheet
        row_count = bulk_generator._create_line_items_sheet(
            sheet,
            sample_associations,
            sample_campaign
        )
        
        # Assertions
        assert row_count == len(sample_associations) + 1  # Header + line item rows
        assert sheet.cell(row=1, column=1).value == "Line Item ID"
        assert sheet.cell(row=2, column=1).value.startswith("camp_123_LI_")
        assert sheet.cell(row=2, column=3).value == "camp_123"
    
    def test_create_targeting_sheet(
        self,
        bulk_generator,
        sample_associations,
        sample_campaign
    ):
        """Test targeting sheet creation."""
        from openpyxl import Workbook
        
        wb = Workbook()
        sheet = wb.active
        
        # Create targeting sheet
        row_count = bulk_generator._create_targeting_sheet(
            sheet,
            sample_associations,
            sample_campaign
        )
        
        # Assertions
        # Should have header + default targeting rules + custom targeting for each line item
        expected_rows = 1 + (6 + 4) * len(sample_associations)  # Default + custom per line item
        assert row_count == expected_rows
        assert sheet.cell(row=1, column=1).value == "Line Item ID"
        assert sheet.cell(row=2, column=2).value == "geo"
        assert sheet.cell(row=2, column=3).value == "US"
    
    def test_create_campaign_dataframe(
        self,
        bulk_generator,
        sample_campaign
    ):
        """Test campaign info DataFrame creation."""
        df = bulk_generator._create_campaign_dataframe(sample_campaign)
        
        # Assertions
        assert len(df) > 0
        assert "Field" in df.columns
        assert "Value" in df.columns
        assert "Campaign ID" in df["Field"].values
        assert "camp_123" in df["Value"].values
    
    def test_create_creatives_dataframe(
        self,
        bulk_generator,
        sample_creatives
    ):
        """Test creatives DataFrame creation."""
        df = bulk_generator._create_creatives_dataframe(sample_creatives, "phase_1")
        
        # Assertions
        assert len(df) == len(sample_creatives)
        assert "Creative ID" in df.columns
        assert "Creative Name" in df.columns
        assert "Phase" in df.columns
        assert all(df["Phase"] == "phase_1")
        assert "creative_1" in df["Creative ID"].values
    
    def test_create_line_items_dataframe(
        self,
        bulk_generator,
        sample_associations,
        sample_campaign
    ):
        """Test line items DataFrame creation."""
        df = bulk_generator._create_line_items_dataframe(sample_associations, sample_campaign)
        
        # Assertions
        assert len(df) == len(sample_associations)
        assert "Line Item ID" in df.columns
        assert "Campaign ID" in df.columns
        assert all(df["Campaign ID"] == "camp_123")
        assert all(df["Line Item ID"].str.startswith("camp_123_LI_"))
    
    @pytest.mark.asyncio
    async def test_download_bulk_sheet(
        self,
        bulk_generator,
        temp_output_dir
    ):
        """Test bulk sheet download."""
        # Create a test file
        test_file_path = os.path.join(temp_output_dir, "test_bulk_sheet.xlsx")
        test_content = b"Test Excel content"
        
        with open(test_file_path, "wb") as f:
            f.write(test_content)
        
        # Download bulk sheet
        content = await bulk_generator.download_bulk_sheet(test_file_path)
        
        # Assertions
        assert content == test_content
    
    @pytest.mark.asyncio
    async def test_download_bulk_sheet_not_found(
        self,
        bulk_generator
    ):
        """Test bulk sheet download with missing file."""
        with pytest.raises(FileNotFoundError):
            await bulk_generator.download_bulk_sheet("/nonexistent/path.xlsx")
    
    @pytest.mark.asyncio
    async def test_list_bulk_sheets(
        self,
        bulk_generator,
        temp_output_dir
    ):
        """Test listing bulk sheets."""
        # Create test files
        test_files = ["bulk_sheet_camp_1.xlsx", "bulk_sheet_camp_2.xlsx"]
        for file_name in test_files:
            file_path = os.path.join(temp_output_dir, file_name)
            with open(file_path, "w") as f:
                f.write("test content")
        
        # List all bulk sheets
        sheets = await bulk_generator.list_bulk_sheets()
        
        # Assertions
        assert len(sheets) == 2
        assert all("file_name" in sheet for sheet in sheets)
        assert all("created_at" in sheet for sheet in sheets)
        assert all("size" in sheet for sheet in sheets)
        
        # List with campaign filter
        filtered_sheets = await bulk_generator.list_bulk_sheets("camp_1")
        assert len(filtered_sheets) == 1
        assert "camp_1" in filtered_sheets[0]["file_name"]