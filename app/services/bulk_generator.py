"""Bulk sheet generator for Amazon DSP campaign activation."""
import io
import os
from typing import Dict, List, Optional, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.database import (
    CampaignDB,
    ProcessedCreativeDB,
    CampaignCreativeAssociationDB
)
from app.models.creative import ViewabilityPhase
from app.utils.logging import get_logger

logger = get_logger("bulk_generator")


class BulkSheetRequest(BaseModel):
    """Request model for bulk sheet generation."""
    campaign_id: str
    include_creatives: bool = True
    include_line_items: bool = True
    include_targeting: bool = True
    format: str = "xlsx"  # xlsx or csv


class BulkSheetResponse(BaseModel):
    """Response model for bulk sheet generation."""
    file_path: str
    file_name: str
    total_rows: int
    sheets: List[str]
    created_at: datetime


class CreativeRow(BaseModel):
    """Model for creative row in bulk sheet."""
    creative_id: str
    creative_name: str
    creative_type: str
    dimensions: str
    click_url: str
    viewability_vendor: str
    phase: str
    status: str = "ACTIVE"
    amazon_creative_id: Optional[str] = None


class LineItemRow(BaseModel):
    """Model for line item row in bulk sheet."""
    line_item_id: str
    line_item_name: str
    campaign_id: str
    creative_ids: str
    budget: float
    bid: float
    bid_type: str = "CPM"
    status: str = "ACTIVE"


class TargetingRow(BaseModel):
    """Model for targeting row in bulk sheet."""
    line_item_id: str
    targeting_type: str
    targeting_value: str
    include_exclude: str = "include"


class BulkSheetGenerator:
    """Generates Excel/CSV bulk sheets for Amazon DSP campaign activation."""
    
    def __init__(self, db_session: AsyncSession, output_dir: str = "/tmp/bulk_sheets"):
        self.db = db_session
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.logger = logger
    
    async def generate_bulk_sheet(
        self,
        request: BulkSheetRequest
    ) -> BulkSheetResponse:
        """Generate bulk sheet for campaign activation."""
        self.logger.info(f"Generating bulk sheet for campaign: {request.campaign_id}")
        
        # Fetch campaign data
        campaign_data = await self._fetch_campaign_data(request.campaign_id)
        
        if not campaign_data:
            raise ValueError(f"Campaign not found: {request.campaign_id}")
        
        # Generate file name
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_name = f"bulk_sheet_{campaign_data['campaign'].campaign_id}_{timestamp}.{request.format}"
        file_path = os.path.join(self.output_dir, file_name)
        
        if request.format == "xlsx":
            response = await self._generate_excel_sheet(
                campaign_data,
                file_path,
                request
            )
        else:
            response = await self._generate_csv_sheet(
                campaign_data,
                file_path,
                request
            )
        
        response.file_path = file_path
        response.file_name = file_name
        response.created_at = datetime.utcnow()
        
        # Update campaign with bulk sheet path
        campaign_data['campaign'].bulk_sheet_path = file_path
        await self.db.commit()
        
        self.logger.info(f"Bulk sheet generated successfully: {file_path}")
        
        return response
    
    async def _fetch_campaign_data(self, campaign_id: str) -> Optional[Dict[str, Any]]:
        """Fetch all campaign data needed for bulk sheet generation."""
        # Fetch campaign
        campaign_result = await self.db.execute(
            select(CampaignDB).where(CampaignDB.campaign_id == campaign_id)
        )
        campaign = campaign_result.scalar_one_or_none()
        
        if not campaign:
            return None
        
        # Fetch creative associations
        associations_result = await self.db.execute(
            select(CampaignCreativeAssociationDB).where(
                CampaignCreativeAssociationDB.campaign_id == campaign_id
            )
        )
        associations = associations_result.scalars().all()
        
        # Fetch creatives
        creative_ids = [assoc.creative_id for assoc in associations]
        creatives_result = await self.db.execute(
            select(ProcessedCreativeDB).where(
                ProcessedCreativeDB.creative_id.in_(creative_ids)
            )
        )
        creatives = creatives_result.scalars().all()
        
        return {
            "campaign": campaign,
            "associations": associations,
            "creatives": creatives
        }
    
    async def _generate_excel_sheet(
        self,
        campaign_data: Dict[str, Any],
        file_path: str,
        request: BulkSheetRequest
    ) -> BulkSheetResponse:
        """Generate Excel bulk sheet."""
        wb = Workbook()
        sheets_created = []
        total_rows = 0
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Campaign Info Sheet
        campaign_sheet = wb.create_sheet("Campaign_Info")
        sheets_created.append("Campaign_Info")
        campaign_rows = self._create_campaign_info_sheet(
            campaign_sheet,
            campaign_data["campaign"]
        )
        total_rows += campaign_rows
        
        # Creatives Sheet
        if request.include_creatives:
            creative_sheet = wb.create_sheet("Creatives")
            sheets_created.append("Creatives")
            creative_rows = self._create_creatives_sheet(
                creative_sheet,
                campaign_data["creatives"],
                campaign_data["campaign"].phase
            )
            total_rows += creative_rows
        
        # Line Items Sheet
        if request.include_line_items:
            line_item_sheet = wb.create_sheet("Line_Items")
            sheets_created.append("Line_Items")
            line_item_rows = self._create_line_items_sheet(
                line_item_sheet,
                campaign_data["associations"],
                campaign_data["campaign"]
            )
            total_rows += line_item_rows
        
        # Targeting Sheet
        if request.include_targeting:
            targeting_sheet = wb.create_sheet("Targeting")
            sheets_created.append("Targeting")
            targeting_rows = self._create_targeting_sheet(
                targeting_sheet,
                campaign_data["associations"],
                campaign_data["campaign"]
            )
            total_rows += targeting_rows
        
        # Save workbook
        wb.save(file_path)
        
        return BulkSheetResponse(
            file_path=file_path,
            file_name=os.path.basename(file_path),
            total_rows=total_rows,
            sheets=sheets_created,
            created_at=datetime.utcnow()
        )
    
    async def _generate_csv_sheet(
        self,
        campaign_data: Dict[str, Any],
        file_path: str,
        request: BulkSheetRequest
    ) -> BulkSheetResponse:
        """Generate CSV bulk sheet."""
        dfs = []
        total_rows = 0
        
        # Campaign Info
        campaign_df = self._create_campaign_dataframe(campaign_data["campaign"])
        dfs.append(("campaign_info", campaign_df))
        total_rows += len(campaign_df)
        
        # Creatives
        if request.include_creatives:
            creatives_df = self._create_creatives_dataframe(
                campaign_data["creatives"],
                campaign_data["campaign"].phase
            )
            dfs.append(("creatives", creatives_df))
            total_rows += len(creatives_df)
        
        # Line Items
        if request.include_line_items:
            line_items_df = self._create_line_items_dataframe(
                campaign_data["associations"],
                campaign_data["campaign"]
            )
            dfs.append(("line_items", line_items_df))
            total_rows += len(line_items_df)
        
        # Save all dataframes to separate CSV files
        base_path = file_path.replace(".csv", "")
        sheets = []
        
        for sheet_name, df in dfs:
            sheet_path = f"{base_path}_{sheet_name}.csv"
            df.to_csv(sheet_path, index=False)
            sheets.append(sheet_name)
        
        return BulkSheetResponse(
            file_path=base_path,
            file_name=os.path.basename(base_path),
            total_rows=total_rows,
            sheets=sheets,
            created_at=datetime.utcnow()
        )
    
    def _create_campaign_info_sheet(self, sheet, campaign: CampaignDB) -> int:
        """Create campaign info sheet."""
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Field", "Value"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        campaign_info = [
            ("Campaign ID", campaign.campaign_id),
            ("Campaign Name", campaign.name),
            ("Advertiser ID", campaign.advertiser_id),
            ("Status", campaign.status),
            ("Phase", campaign.phase),
            ("Budget", f"${campaign.total_budget:,.2f}"),
            ("Start Date", campaign.start_date.strftime("%Y-%m-%d")),
            ("End Date", campaign.end_date.strftime("%Y-%m-%d")),
            ("Creative Count", campaign.creative_count),
            ("Processed Creatives", campaign.processed_creatives_count),
            ("Amazon Order ID", campaign.order_id or "N/A"),
            ("Created At", campaign.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Updated At", campaign.updated_at.strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for row_idx, (field, value) in enumerate(campaign_info, 2):
            sheet.cell(row=row_idx, column=1, value=field).border = border
            sheet.cell(row=row_idx, column=2, value=value).border = border
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40
        
        return len(campaign_info) + 1
    
    def _create_creatives_sheet(
        self,
        sheet,
        creatives: List[ProcessedCreativeDB],
        phase: str
    ) -> int:
        """Create creatives sheet."""
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Creative ID", "Creative Name", "Format", "Type", "Dimensions",
            "Viewability Phase", "Viewability Vendor", "Amazon Creative ID",
            "Status", "Amazon DSP Ready"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        for row_idx, creative in enumerate(creatives, 2):
            viewability_config = creative.viewability_config or {}
            vendors = viewability_config.get("vendors", [])
            vendor_str = ", ".join(vendors) if vendors else "N/A"
            
            # Extract dimensions from processing metadata
            processing_metadata = creative.processing_metadata or {}
            dimensions = processing_metadata.get("dimensions", "N/A")
            
            row_data = [
                creative.creative_id,
                creative.name,
                creative.format,
                creative.creative_type,
                dimensions,
                phase,
                vendor_str,
                creative.amazon_creative_id or "Pending",
                creative.status.upper(),
                "Yes" if creative.amazon_dsp_ready else "No"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 15
        
        return len(creatives) + 1
    
    def _create_line_items_sheet(
        self,
        sheet,
        associations: List[CampaignCreativeAssociationDB],
        campaign: CampaignDB
    ) -> int:
        """Create line items sheet."""
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Line Item ID", "Line Item Name", "Campaign ID", "Creative ID",
            "Type", "Budget", "Bid", "Bid Type", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        for row_idx, assoc in enumerate(associations, 2):
            line_item_id = f"{campaign.campaign_id}_LI_{row_idx-1:03d}"
            
            row_data = [
                line_item_id,
                assoc.line_item_name,
                assoc.campaign_id,
                assoc.creative_id,
                assoc.line_item_type,
                f"${assoc.budget:,.2f}",
                f"${assoc.bid:.2f}",
                "CPM",
                assoc.status.upper()
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 15
        
        return len(associations) + 1
    
    def _create_targeting_sheet(
        self,
        sheet,
        associations: List[CampaignCreativeAssociationDB],
        campaign: CampaignDB
    ) -> int:
        """Create targeting sheet."""
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Line Item ID", "Targeting Type", "Targeting Value", "Include/Exclude"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Extract targeting from campaign config
        config = campaign.config or {}
        targeting = config.get("targeting", {})
        
        row_idx = 2
        
        # Default targeting rules
        default_targeting = [
            ("geo", "US", "include"),
            ("device", "desktop", "include"),
            ("device", "mobile", "include"),
            ("device", "tablet", "include"),
            ("viewability", "70%+", "include"),
            ("brand_safety", "low_risk", "include")
        ]
        
        # Add custom targeting if available
        if targeting:
            if "audiences" in targeting:
                for audience in targeting.get("audiences", []):
                    default_targeting.append(("audience", audience, "include"))
            if "keywords" in targeting:
                for keyword in targeting.get("keywords", []):
                    default_targeting.append(("keyword", keyword, "include"))
        
        # Generate targeting rows for each line item
        for assoc_idx, assoc in enumerate(associations):
            line_item_id = f"{campaign.campaign_id}_LI_{assoc_idx+1:03d}"
            
            for targeting_type, targeting_value, include_exclude in default_targeting:
                row_data = [
                    line_item_id,
                    targeting_type,
                    targeting_value,
                    include_exclude
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                
                row_idx += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 20
        
        return row_idx - 1
    
    def _create_campaign_dataframe(self, campaign: CampaignDB) -> pd.DataFrame:
        """Create campaign info dataframe."""
        return pd.DataFrame({
            "Field": [
                "Campaign ID", "Campaign Name", "Advertiser ID", "Status",
                "Phase", "Budget", "Start Date", "End Date",
                "Creative Count", "Amazon Order ID"
            ],
            "Value": [
                campaign.campaign_id,
                campaign.name,
                campaign.advertiser_id,
                campaign.status,
                campaign.phase,
                campaign.total_budget,
                campaign.start_date.strftime("%Y-%m-%d"),
                campaign.end_date.strftime("%Y-%m-%d"),
                campaign.creative_count,
                campaign.order_id or "N/A"
            ]
        })
    
    def _create_creatives_dataframe(
        self,
        creatives: List[ProcessedCreativeDB],
        phase: str
    ) -> pd.DataFrame:
        """Create creatives dataframe."""
        data = []
        for creative in creatives:
            viewability_config = creative.viewability_config or {}
            vendors = viewability_config.get("vendors", [])
            vendor_str = ", ".join(vendors) if vendors else "N/A"
            
            data.append({
                "Creative ID": creative.creative_id,
                "Creative Name": creative.name,
                "Format": creative.format,
                "Type": creative.creative_type,
                "Phase": phase,
                "Viewability Vendor": vendor_str,
                "Amazon Creative ID": creative.amazon_creative_id or "Pending",
                "Status": creative.status.upper()
            })
        
        return pd.DataFrame(data)
    
    def _create_line_items_dataframe(
        self,
        associations: List[CampaignCreativeAssociationDB],
        campaign: CampaignDB
    ) -> pd.DataFrame:
        """Create line items dataframe."""
        data = []
        for idx, assoc in enumerate(associations):
            line_item_id = f"{campaign.campaign_id}_LI_{idx+1:03d}"
            
            data.append({
                "Line Item ID": line_item_id,
                "Line Item Name": assoc.line_item_name,
                "Campaign ID": assoc.campaign_id,
                "Creative ID": assoc.creative_id,
                "Type": assoc.line_item_type,
                "Budget": assoc.budget,
                "Bid": assoc.bid,
                "Status": assoc.status.upper()
            })
        
        return pd.DataFrame(data)
    
    async def download_bulk_sheet(self, file_path: str) -> bytes:
        """Download bulk sheet as bytes for API response."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Bulk sheet not found: {file_path}")
        
        with open(file_path, "rb") as f:
            return f.read()
    
    async def list_bulk_sheets(self, campaign_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """List available bulk sheets."""
        sheets = []
        
        for file_name in os.listdir(self.output_dir):
            if campaign_id and campaign_id not in file_name:
                continue
            
            file_path = os.path.join(self.output_dir, file_name)
            stat = os.stat(file_path)
            
            sheets.append({
                "file_name": file_name,
                "file_path": file_path,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_ctime),
                "modified_at": datetime.fromtimestamp(stat.st_mtime)
            })
        
        return sorted(sheets, key=lambda x: x["created_at"], reverse=True)